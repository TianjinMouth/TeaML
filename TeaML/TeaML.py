from .utils.tea_utils import *
from .utils.tea_filter import *
from .utils.auto_bin_woe import AutoBinWOE
from .utils.tea_encoder import *
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import StratifiedKFold
from openpyxl.chart import BarChart, Reference, LineChart
import warnings
import configparser
import os
current_file_dir = os.path.dirname(__file__)
conf = configparser.ConfigParser()

conf.read(current_file_dir + '/conf.ini', encoding='utf-8')
warnings.filterwarnings('ignore')


def woe_transformer(df, woe_dict):
    _df = df.copy()
    for f in woe_dict:
        for threshold in woe_dict[f]:
            woe_value = woe_dict[f][threshold]
            threshold = threshold.replace('[', '').replace(' ', '').replace(')', '')
            left, right = threshold.split(',')
            if left == 'nan':
                _df.loc[df[f].isnull(), f] = woe_value
            else:
                _df.loc[(df[f] >= float(left)) & (df[f] < float(right)), f] = woe_value
    return _df


def woe_todict(sheet_feature_bin_ins):
    woe_dict = {}
    for f in [i for i in sheet_feature_bin_ins['feature'].unique()]:
        inner_map = {}
        _tmp = sheet_feature_bin_ins[sheet_feature_bin_ins['feature']==f]
        for index, row in _tmp.iterrows():
            inner_map['[%.6f, %.6f)' %(row.left, row.right)] = np.round(row.woe, 6)
        woe_dict[f] = inner_map
    return woe_dict


class WOE:
    def __init__(self, bins=10, psi_threshold=None, monotony_merge=True, bad_rate_merge=False,
                 bad_rate_sim_threshold=0.05, chi2_merge=False, chi2_threshold=3.841, iv_threshold=None,
                 prune=False, prune_threshold=0.05):
        self.bins = bins
        self.psi_threshold = psi_threshold
        self.bad_rate_merge = bad_rate_merge
        self.bad_rate_sim_threshold = bad_rate_sim_threshold
        self.chi2_merge = chi2_merge
        self.chi2_threshold = chi2_threshold
        self.monotony_merge = monotony_merge
        self.iv_threshold = iv_threshold
        self.prune = prune
        self.prune_threshold = prune_threshold

    def woe_processing(self, x_train, y_train, x_oot, y_oot, gating=True):
        # WOEÁºñÁ†Å
        woe = AutoBinWOE(bins=self.bins, monotony_merge=self.monotony_merge, bad_rate_merge=self.bad_rate_merge,
                         bad_rate_sim_threshold=self.bad_rate_sim_threshold,
                         chi2_merge=self.chi2_merge, chi2_threshold=self.chi2_threshold,
                         prune=self.prune, prune_threshold=self.prune_threshold)
        woe.fit(x_train, y_train)
        x_woe = woe.transform(x_train)
        x_oot_woe = woe.transform(x_oot)

        # ÂèòÈáèÁ®≥ÂÆöÊÄß
        train_bin = woe.cal_bin_ks(x_train[x_woe.columns], y_train)
        oot_bin = woe.cal_bin_ks(x_oot[x_woe.columns], y_oot, oot=True)

        # KS
        ks_ins = []
        ks_oot = []
        ks_index = []
        for i in x_woe.columns:
            try:
                ks_ins.append(max(train_bin[i][train_bin[i].ks.notnull()].ks))
                ks_oot.append(max(oot_bin[i][oot_bin[i].ks.notnull()].ks))
                ks_index.append(i)
            except Exception:
                pass

        # IV
        iv_data = []
        iv_index = []
        for i in x_woe.columns:
            try:
                tmp = sum(woe.data_matrix[i].iv)
                if tmp == 0:
                    pass
                else:
                    iv_index.append(i)
                    iv_data.append(tmp)
            except Exception:
                pass

        # PSI
        psi = pd.DataFrame(woe.cal_psi(train_bin, oot_bin), index=['psi']).T.reset_index().rename(
            columns={'index': 'Â≠óÊÆµÂêçÁß∞'})
        iv = pd.DataFrame({'Â≠óÊÆµÂêçÁß∞': iv_index, 'Information Value': iv_data})
        ks = pd.DataFrame({'Â≠óÊÆµÂêçÁß∞': ks_index, 'INS_KS': ks_ins, 'OOT_KS': ks_oot})
        psi_ks_iv = psi.merge(iv, how='inner', on='Â≠óÊÆµÂêçÁß∞').merge(ks, how='inner', on='Â≠óÊÆµÂêçÁß∞')

        if self.psi_threshold is None:
            pass
        else:
            if isinstance(self.psi_threshold, float):
                left_features = psi_ks_iv[psi_ks_iv['psi'] < self.psi_threshold]['Â≠óÊÆµÂêçÁß∞'].tolist()
                psi_ks_iv = psi_ks_iv[psi_ks_iv['psi'] < self.psi_threshold]
                x_woe = x_woe[left_features]
                x_oot_woe = x_oot_woe[left_features]
            else:
                raise ValueError("psi_threshold must be 'all' or a float between (0, 1]")

        if self.iv_threshold is None:
            pass
        else:
            if isinstance(self.iv_threshold, float):
                left_features = psi_ks_iv[psi_ks_iv['Information Value'] >= self.iv_threshold]['Â≠óÊÆµÂêçÁß∞'].tolist()
                psi_ks_iv = psi_ks_iv[psi_ks_iv['Information Value'] >= self.iv_threshold]
                x_woe = x_woe[left_features]
                x_oot_woe = x_oot_woe[left_features]
            else:
                raise ValueError("iv_threshold must be a float like 0.01")

        if gating:
            return x_woe, x_oot_woe, woe, psi_ks_iv
        else:
            return x_train, x_oot, woe, psi_ks_iv


class Tea:
    def __init__(self, useless_features, label='is_overdue', datetime_feature='create_time', split_method='oot'
                 , oot_start=None, oot_end=None, file_path='final_report.xlsx', embellish=True):
        """

        :param useless_features: list
        :param label: str
        :param datetime_feature: str
        :param split_method: 'oos' or 'oot'
        :param oot_start: if split_method == 'oot'
        :param oot_end: if split_method == 'oot'
        :param file_path: file_path
        """
        self.oot_start = oot_start
        self.oot_end = oot_end
        self.sheets = dict()
        self.ct = None  # ÂàÜÁ±ªÂèòÈáèbad rateÊõøÊç¢Á±ª
        self.left_features = []  # ÂÖ•Ê®°ÂûãÁöÑÂèòÈáè
        self.woe = None  # woeÂàÜÁÆ±Á±ª
        self.clf = None  # ÂàÜÁ±ªÊ®°Âûã
        self.X_train = None  # ÂéüÂßãËÆ≠ÁªÉÊï∞ÊçÆ
        self.X_oot = None  # ÂéüÂßãOOTÊï∞ÊçÆ
        self.y_train = None  # ËÆ≠ÁªÉÁöÑY
        self.y_oot = None  # OOTÁöÑY
        self.stacking_train = None  # Ê®°ÂûãcvÁöÑËÆ≠ÁªÉÊ¶ÇÁéáÂàÜ
        self.stacking_oot = None  # Ê®°ÂûãcvÁöÑOOTÊ¶ÇÁéáÂàÜ
        self.train_ts = None  # ËÆ≠ÁªÉÊï∞ÊçÆÁöÑÊó∂Èó¥Â∫èÂàó
        self.oot_ts = None  # OOTÊï∞ÊçÆÁöÑÊó∂Èó¥Â∫èÂàó
        self.datetime_feature = datetime_feature  # Êó∂Èó¥ÂàóÁöÑÂàóÂêç
        self.split_method = split_method
        self.label = label  # yÂÄºÂàóÁöÑÂàóÂêç
        self.useless_features = useless_features  # ‰∏çÂèÇ‰∏éËÆ≠ÁªÉÁöÑÂàó
        self.file_path = file_path  # Êä•Âëä‰øùÂ≠òÂú∞ÂùÄ
        self.embellish = embellish  # ÊòØÂê¶ËøõË°å‰øÆÈ•∞
        self.method = None  # ÂèòÈáèÁ≠õÈÄâÁöÑÈ°∫Â∫èÂàóË°®
        self.encoders = None  # ÁºñÁ†ÅÂô®
        self.bins = None
        self.method_judge = []

    @staticmethod
    def get_describe(x):
        nu = []
        nu_ratio = []
        most_common = []
        most_common_ratio = []
        for i in x.columns:
            if len(x[i].value_counts()) == 0:
                most_common.append(len(x[i]))
                most_common_ratio.append(1.0)
            else:
                most_common.append(sum(x[i] == (x[i].value_counts().index[0])))
                most_common_ratio.append(sum(x[i] == (x[i].value_counts().index[0])) / x.shape[0])
            nu.append(sum(x[i].isnull()))
            nu_ratio.append(sum(x[i].isnull()) / x.shape[0])
        return nu, nu_ratio, most_common, most_common_ratio

    @staticmethod
    def _ks_curve(df, month=None):
        a = pd.DataFrame()
        a['decile'] = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
        a['bin'] = df.iloc[:, 0]
        a['size'] = df.iloc[:, 1]
        a['num_bad'] = df.iloc[:, 2]
        a['prop_bad'] = a['num_bad'] / sum(a['num_bad'])
        a['cum_prop_bad'] = a['prop_bad'][0]
        for i in range(1, a.shape[0]):
            a['cum_prop_bad'][i] = a['cum_prop_bad'][i - 1] + a['prop_bad'][i]
        a['num_good'] = a['size'] - a['num_bad']
        a['prop_good'] = a['num_good'] / sum(a['num_good'])
        a['cum_prop_good'] = a['prop_good'][0]
        for i in range(1, a.shape[0]):
            a['cum_prop_good'][i] = a['cum_prop_good'][i - 1] + a['prop_good'][i]
        a['bad_rate'] = round(a['num_bad'] / a['size'], 4)
        a['ks_curve'] = a['cum_prop_good'] - a['cum_prop_bad']
        a['ks_curve'] = round(a['ks_curve'], 4)
        if month is not None:
            a.insert(0, 'month', month)
        return a

    def wash(self, data, null_drop_rate=0.8, most_common_drop_rate=0.9, init_data=False, train_set=None, oot_set=None):
        # ------------------------  STEP 1 ËÆ≠ÁªÉÊµãËØïÈõÜÂàíÂàÜ„ÄÅBad rateÊõøÊç¢ÂèäÂèòÈáèÂàùÁ≠õ--------------------------------
        df = data.copy()
        df[self.datetime_feature] = pd.to_datetime(df[self.datetime_feature])

        if self.split_method == 'oot':
            oot = df[(df[self.datetime_feature] >= self.oot_start) & (df[self.datetime_feature] < self.oot_end)]
            train = df.drop(oot.index).reset_index(drop=True)
            oot = oot.reset_index(drop=True)

        elif self.split_method == 'oos':
            from sklearn.model_selection import train_test_split
            train, oot, y_train, y_oot = train_test_split(df, df[self.label], test_size=0.2, random_state=12)
            train = train.reset_index(drop=True)
            oot = oot.reset_index(drop=True)
            y_train = y_train.reset_index(drop=True)
            y_oot = y_oot.reset_index(drop=True)
        else:
            raise KeyError("Wrong type...")

        if init_data:
            self.X_train = train.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
            self.y_train = train[self.label]
            self.X_oot = oot.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
            self.y_oot = oot[self.label]
            self.train_ts = train_set[self.datetime_feature]
            self.oot_ts = oot_set[self.datetime_feature]
            self.sheets['sheet_distribution'] = None
        else:
            X = df.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
            X_train = train.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
            y_train = train[self.label]
            X_oot = oot.drop(self.useless_features + [self.datetime_feature] + [self.label], axis=1)
            y_oot = oot[self.label]
            self.train_ts = train[self.datetime_feature]
            self.oot_ts = oot[self.datetime_feature]

            # ==== sheet Ê†∑Êú¨ÂàÜÊûê ====
            sheet_sample = pd.DataFrame(
                {'Êó∂Èó¥Ë∑®Â∫¶': [str(min(train[self.datetime_feature])) + '~' + str(max(train[self.datetime_feature])),
                          str(min(oot[self.datetime_feature])) + '~' + str(max(oot[self.datetime_feature]))],
                 'Cnt': [X_train.shape[0], X_oot.shape[0]],
                 'GoodCnt': [train.shape[0] - sum(train[self.label]), oot.shape[0] - sum(oot[self.label])],
                 'BadCnt': [sum(train[self.label]), sum(oot[self.label])],
                 'BadRate': [sum(y_train) / y_train.shape[0], sum(y_oot) / y_oot.shape[0]]})
            self.sheets['sheet_sample'] = sheet_sample
            # ==== sheet ÂèòÈáèÁº∫Â§±Áéá & Âü∫Êú¨Êé¢Á¥¢ÊÄßÂàÜÊûê ====
            nu, nu_ratio, most_common, most_common_ratio = Tea.get_describe(X)

            # ÂèòÈáèÂàùÁ≠õ
            print("Preliminary screening...")
            sheet_2_tmp = pd.merge(pd.DataFrame(
                {'ÂèòÈáèÂêçÁß∞': list(X.columns), 'Á©∫ÂÄº‰∏™Êï∞': nu, 'Á©∫ÂÄº‰∏™Êï∞Âç†ÊØî': nu_ratio, 'ÊúÄÂ∏∏ÂÄº‰∏™Êï∞': most_common, 'ÊúÄÂ∏∏ÂÄº‰∏™Êï∞Âç†ÊØî': most_common_ratio}),
                pd.DataFrame(
                    X.describe().T.reset_index()[['index', 'mean', 'std', 'min', '25%', '50%', '75%', 'max']]).rename(
                    columns={'index': 'ÂèòÈáèÂêçÁß∞'}),
                how='left', on='ÂèòÈáèÂêçÁß∞')
            left_features = list(set(sheet_2_tmp[sheet_2_tmp['Á©∫ÂÄº‰∏™Êï∞Âç†ÊØî'] < null_drop_rate]['ÂèòÈáèÂêçÁß∞']) & set(
                sheet_2_tmp[sheet_2_tmp['ÊúÄÂ∏∏ÂÄº‰∏™Êï∞Âç†ÊØî'] < most_common_drop_rate]['ÂèòÈáèÂêçÁß∞']))

            X = X[left_features]
            X_train = X_train[left_features]
            X_oot = X_oot[left_features]

            nu, nu_ratio, most_common, most_common_ratio = self.get_describe(X)

            sheet_distribution = pd.concat([pd.DataFrame(
                {'ÂèòÈáèÂêçÁß∞': list(X.columns), 'Á©∫ÂÄº‰∏™Êï∞': nu, 'Á©∫ÂÄº‰∏™Êï∞Âç†ÊØî': nu_ratio, 'ÊúÄÂ∏∏ÂÄº‰∏™Êï∞': most_common, 'ÊúÄÂ∏∏ÂÄº‰∏™Êï∞Âç†ÊØî': most_common_ratio}),
                pd.DataFrame(X.describe().T.reset_index()[
                                 ['mean', 'std', 'min', '25%', '50%', '75%', 'max']])], axis=1)

            self.sheets['sheet_distribution'] = sheet_distribution
            self.X_train = X_train
            self.X_oot = X_oot
            self.y_train = y_train
            self.y_oot = y_oot

    def cook(self, encoders):
        self.encoders = encoders
        for estimator in self.encoders:
            if isinstance(estimator, TeaOneHotEncoder):
                self.X_train = estimator.fit_transform(self.X_train)
                self.X_oot = estimator.transform(self.X_oot)
            else:
                self.X_train = estimator.fit_transform(self.X_train, self.y_train)
                self.X_oot = estimator.transform(self.X_oot)

    def select(self, method):
        self.method = method
        _X = self.X_train.copy()
        _X_oot = self.X_oot.copy()

        try:
            self.method_judge = [1 if isinstance(estimator, WOE) else 0 for estimator in method]
        except:
            pass

        if sum(self.method_judge) != 0:
            for estimator in self.method:
                if isinstance(estimator, WOE):
                    _X, _X_oot, woe, psi_ks_iv = estimator.woe_processing(_X, self.y_train, _X_oot, self.y_oot, gating=True)
                elif isinstance(estimator, OutlierTransform) \
                        or isinstance(estimator, FilterCorr) \
                        or isinstance(estimator, FilterVif):
                    _X = estimator.fit_transform(_X)
                    _X_oot = estimator.transform(_X_oot)
                else:
                    _X = estimator.fit_transform(_X, self.y_train)
                    _X_oot = estimator.transform(_X_oot)
        elif sum(self.method_judge) == 0:
            for estimator in self.method:
                if isinstance(estimator, OutlierTransform) \
                        or isinstance(estimator, FilterCorr) \
                        or isinstance(estimator, FilterVif):
                    _X = estimator.fit_transform(_X)
                    _X_oot = estimator.transform(_X_oot)
                else:
                    _X = estimator.fit_transform(_X, self.y_train)
                    _X_oot = estimator.transform(_X_oot)
            wow = WOE(bins=10, bad_rate_merge=True, bad_rate_sim_threshold=0.05)
            _X, _X_oot, woe, psi_ks_iv = wow.woe_processing(_X, self.y_train, _X_oot, self.y_oot, gating=False)

        X_woe, X_oot_woe = _X, _X_oot
        self.left_features = X_woe.columns.tolist()
        sheet_psi_ks_iv = psi_ks_iv[psi_ks_iv['Â≠óÊÆµÂêçÁß∞'].isin(self.left_features)].reset_index(drop=True)
        self.sheets['sheet_psi_ks_iv'] = sheet_psi_ks_iv
        self.woe = woe
        self.X_woe = X_woe
        self.X_oot_woe = X_oot_woe

    def drink(self, clf, cross_validation):
        if self.method_judge != 0:
            self.X_oot_woe = self.X_oot_woe.fillna(0.0)
        else:
            pass
        clf, stacking_train, stacking_oot = train_by_cv(self.X_woe, self.y_train, self.X_oot_woe, self.y_oot,
                                                        cross_validation, clf)
        if judge_h2o(clf):
            full_training_frame = h2o.H2OFrame(pd.concat([self.X_woe, self.y_train], axis=1))
            full_training_frame[self.y_train.name] = full_training_frame[self.y_train.name].asfactor()
            clf.train(x=list(self.X_woe.columns), y=self.y_train.name, training_frame=full_training_frame)
        else:
            clf.fit(self.X_woe, self.y_train)
        self.clf = clf
        self.stacking_train = stacking_train
        self.stacking_oot = stacking_oot

    def sleep(self, bins):
        # ------------------------------  STEP 4 Êï¥ÁêÜË°®  ---------------------------------------------------------------
        # ==== sheet ÂêÑ‰∏™binÁöÑÂèòÈáèÈÄæÊúüÂàÜÂ∏ÉÂíåKSÂÄº ====
        self.bins = bins
        train_bin = self.woe.cal_bin_ks(self.X_train[self.left_features], self.y_train)
        oot_bin = self.woe.cal_bin_ks(self.X_oot[self.left_features], self.y_oot, oot=True)

        # ËØÑÂàÜÁõ∏ÂÖ≥ÊÄß
        sheet_correlations = self.X_train[self.left_features].corr()

        # ==== sheet ÁâπÂæÅÊùÉÈáçÂíåÈ¢ÑÊµãÊ¶ÇÁéáÂàÜÊûê ====
        # result_train = pd.concat([X_woe, pd.DataFrame(stacking_train)], axis=1).rename(columns={0: 'model_score'})
        result_test = pd.concat([self.X_oot_woe, pd.DataFrame(self.stacking_oot)], axis=1).rename(
            columns={0: 'model_score'})
        if judge_h2o(self.clf):
            im_final = self.clf._model_json['output']['variable_importances'].as_data_frame()[['variable', 'percentage']].rename(columns={'variable': 'feature_name', 'percentage': 'feature_coef / percentage'}).reset_index(drop=True)
        else:
            im_final = get_importance(self.clf, self.X_woe).reset_index(drop=True)

        model_col = list(set(result_test.columns) - {'model_score'})
        score_corr = []
        for i in model_col:
            score_corr.append(result_test[i].corr(result_test['model_score']))

        label_corr = []
        for i in model_col:
            label_corr.append(result_test[i].corr(pd.Series(self.y_oot)))

        tmp = pd.merge(pd.DataFrame({'feature_name': model_col, '‰∏éÈ¢ÑÊµãÊ¶ÇÁéáÁõ∏ÂÖ≥ÊÄß': score_corr}),
                       pd.DataFrame({'feature_name': model_col, '‰∏éÈ£éÈô©Áõ∏ÂÖ≥ÊñπÂêë': label_corr}),
                       how='inner',
                       on='feature_name')

        # ==== sheet Ê®°ÂûãÂàÜÊûêÔºåksÊõ≤Á∫øÔºåÁ¥ØËÆ°bad rate... ====
        sheet_weights = pd.merge(tmp, im_final, how='inner', on='feature_name')
        model_info_tmp_train = pd.concat(
            [self.train_ts, self.y_train, pd.DataFrame(self.stacking_train)],
            axis=1).rename(columns={0: 'model_score'})
        model_info_tmp_oot = pd.concat(
            [self.oot_ts, self.y_oot, pd.DataFrame(self.stacking_oot)],
            axis=1).rename(columns={0: 'model_score'})
        model_info_tmp_train['tag'] = 'INS'
        model_info_tmp_oot['tag'] = 'OOT'
        model_info_tmp = pd.concat([model_info_tmp_train, model_info_tmp_oot])
        model_info_tmp['bin'] = pd.qcut(model_info_tmp['model_score'], 10)
        model_info_tmp['month'] = model_info_tmp[self.datetime_feature].dt.strftime('%Y%m')

        model_info_tmp_ins = pd.DataFrame(model_info_tmp[model_info_tmp['tag'] == 'INS'].groupby('bin').agg(
            {self.datetime_feature: ['count'], self.label: ['sum']})).reset_index()
        model_info_tmp_oot = pd.DataFrame(model_info_tmp[model_info_tmp['tag'] == 'OOT'].groupby('bin').agg(
            {self.datetime_feature: ['count'], self.label: ['sum']})).reset_index()
        sheet_model_info_ins = Tea._ks_curve(model_info_tmp_ins)
        sheet_model_info_ins['tag'] = 'INS'
        sheet_model_info_oot = Tea._ks_curve(model_info_tmp_oot)
        sheet_model_info_oot['tag'] = 'OOT'
        self.sheets['model_info_tmp'] = model_info_tmp

        # -------------------------------  STEP 5 ÂÜôÂÖ•Ë°®  ---------------------------------------------------------------
        bin_ks_ins = pd.DataFrame()
        bin_ks_oot = pd.DataFrame()
        for i in train_bin.keys():
            train_tmp = train_bin[i].reset_index().rename(columns={'index': 'bins'})
            train_tmp.insert(0, 'feature', i)
            bin_ks_ins = pd.concat([bin_ks_ins, train_tmp])
            oot_tmp = oot_bin[i].reset_index().rename(columns={'index': 'bins'})
            oot_tmp.insert(0, 'feature', i)
            bin_ks_oot = pd.concat([bin_ks_oot, oot_tmp])

        self.sheets['sheet_feature_bin_ins'] = bin_ks_ins
        self.sheets['sheet_feature_bin_oot'] = bin_ks_oot
        self.sheets['sheet_correlations'] = sheet_correlations
        self.sheets['sheet_weights'] = sheet_weights
        self.sheets['sheet_model_info_ins'] = sheet_model_info_ins
        self.sheets['sheet_model_info_oot'] = sheet_model_info_oot

        writer = pd.ExcelWriter(self.file_path)
        self.sheets['sheet_sample'].to_excel(writer, sheet_name='Ê†∑Êú¨ÂàÜÊûê', index=False)
        self.sheets['sheet_distribution'].to_excel(writer, sheet_name='ÂèòÈáèÁº∫Â§±Áéá & Âü∫Êú¨Êé¢Á¥¢ÊÄßÂàÜÊûê', index=False)
        self.sheets['sheet_psi_ks_iv'].to_excel(writer, sheet_name='INSÂèòÈáèIVÂÄº & Êó∂Èó¥Á®≥ÂÆöÊÄß', index=False)

        row_index = 0
        for i in train_bin.keys():
            train_tmp = train_bin[i].reset_index().rename(columns={'index': i})
            train_tmp['bad'] = '[' + train_tmp['left'].astype(str) + ',' + train_tmp['right'].astype(str) + ')'
            oot_tmp = oot_bin[i].reset_index().rename(columns={'index': i})
            oot_tmp['bad'] = '[' + oot_tmp['left'].astype(str) + ',' + oot_tmp['right'].astype(str) + ')'
            train_tmp.to_excel(writer, startrow=row_index, startcol=1, sheet_name='ÂèòÈáèÈÄæÊúüÂàÜÂ∏ÉÂíåKSÂÄº', index=False)
            oot_tmp.to_excel(writer, startrow=row_index, startcol=13, sheet_name='ÂèòÈáèÈÄæÊúüÂàÜÂ∏ÉÂíåKSÂÄº', index=False)
            row_index += self.bins + 2

        sheet_correlations.to_excel(writer, sheet_name='ËØÑÂàÜÁõ∏ÂÖ≥ÊÄß', index=True)
        sheet_weights.to_excel(writer, sheet_name='Ê®°Âûã', index=False)
        sheet_model_info_ins.to_excel(writer, sheet_name='Ê®°Âûã', startcol=6, index=False)
        sheet_model_info_oot.to_excel(writer, sheet_name='Ê®°Âûã', startrow=11, startcol=6, index=False, header=False)

        row_index_8 = 0
        for month in model_info_tmp['month'].unique():
            trace_back = model_info_tmp[model_info_tmp['month'] == month].drop(['month'], axis=1).reset_index(drop=True)
            trace_back = pd.DataFrame(
                trace_back.groupby('bin').agg({self.datetime_feature: ['count'], self.label: ['sum']})).reset_index()
            sheet_trace_back = Tea._ks_curve(trace_back, month)
            if row_index_8 == 0:
                sheet_trace_back.to_excel(writer, sheet_name='Ê®°ÂûãÂõûÊµã', startrow=row_index_8, startcol=0, index=False)
                row_index_8 += 11
            else:
                sheet_trace_back.to_excel(writer, sheet_name='Ê®°ÂûãÂõûÊµã', startrow=row_index_8, startcol=0, index=False,
                                          header=False)
                row_index_8 += 10
        writer.save()

        def woe_dump(self):
            if hasattr(self, 'sheets'):
                if 'sheet_feature_bin_ins' in self.sheets.keys():
                    sheet_feature_bin_ins = self.sheets['sheet_feature_bin_ins']
                    woe_dict = {}
                    for f in sheet_feature_bin_ins['feature'].unique():
                        inner_map = {}
                        _tmp = sheet_feature_bin_ins[sheet_feature_bin_ins['feature'] == f]
                        for index, row in _tmp.iterrows():
                            inner_map['[%s, %s)' % (row.left, row.right)] = row.woe
                        woe_dict[f] = inner_map
                    return woe_dict
                else:
                    raise KeyError(" 'sheet_feature_bin_ins' not in tea.sheets!")
            else:
                raise AttributeError("no attribute 'sheets'!")

        # -------------------------------  STEP 6 ÁæéÂåñÔºàÂ≠ó‰Ωì/Â≠óÂè∑/ËæπÊ°Ü/È¢úËâ≤/Á≤óÁªÜÔºâ  --------------------------------------------
        if self.embellish:
            wb = openpyxl.load_workbook(self.file_path)
            left, right, top, bottom = [Side(style='thin', color='000000')] * 4

            sheet = wb['Ê†∑Êú¨ÂàÜÊûê']
            for i in sheet['A1':'E3']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
            for i in sheet['A1':'E1']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            sheet = wb['ÂèòÈáèÁº∫Â§±Áéá & Âü∫Êú¨Êé¢Á¥¢ÊÄßÂàÜÊûê']
            for i in sheet['A1':'L%s' % (self.sheets['sheet_distribution'].shape[0] + 1)]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
            for i in sheet['A1':'L1']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            sheet = wb['INSÂèòÈáèIVÂÄº & Êó∂Èó¥Á®≥ÂÆöÊÄß']
            for i in sheet['A1':'E%s' % (self.sheets['sheet_psi_ks_iv'].shape[0] + 1)]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
            for i in sheet['A1':'E1']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            sheet = wb['ÂèòÈáèÈÄæÊúüÂàÜÂ∏ÉÂíåKSÂÄº']
            for i in sheet['A1':'X%s' % row_index]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].fill = eval(conf.get('config', 'fill2'))
            for ind in range(1, row_index, self.bins+2):
                for i in sheet['B%s' % ind:'L%s' % ind]:
                    for j in range(len(i)):
                        i[j].font = eval(conf.get('config', 'font2'))
                        i[j].fill = eval(conf.get('config', 'fill1'))

                for i in sheet['N%s' % ind:'X%s' % ind]:
                    for j in range(len(i)):
                        i[j].font = eval(conf.get('config', 'font2'))
                        i[j].fill = eval(conf.get('config', 'fill1'))

                sheet['A%s' % ind] = 'INS'
                sheet['M%s' % ind] = 'OOT'
                sheet['A%s' % ind].font = eval(conf.get('config', 'font3'))
                sheet['M%s' % ind].font = eval(conf.get('config', 'font3'))

            for i in sheet['A1':'A%s' % row_index]:
                for j in range(len(i)):
                    i[j].fill = eval(conf.get('config', 'fill3'))

            for i in sheet['M1':'M%s' % row_index]:
                for j in range(len(i)):
                    i[j].fill = eval(conf.get('config', 'fill3'))

            sheet = wb['Ê®°Âûã']
            for i in sheet['A1':'D%s' % (sheet_weights.shape[0] + 1)]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
            for i in sheet['A1':'D1']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            for i in sheet['G1':'R21']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
            for i in sheet['G1':'R1']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            cnt = 2
            while sheet['R'+str(cnt)].value is not None:
                c1 = BarChart()
                v1 = Reference(sheet, min_col=16, min_row=cnt-1, max_col=16, max_row=cnt+9)
                c1.add_data(v1, titles_from_data=True)

                c1.x_axis.title = 'Bin Decile'
                c1.y_axis.title = 'Bad Rate'
                c1.y_axis.majorGridlines = None
                c1.title = sheet['R'+str(cnt)].value

                # Create a second chart
                c2 = LineChart()
                v2 = Reference(sheet, min_col=17, min_row=cnt-1, max_col=17, max_row=cnt+9)
                c2.add_data(v2, titles_from_data=True)
                c2.y_axis.axId = 200
                c2.y_axis.title = "Ks"

                # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
                c1.y_axis.crosses = "max"
                c1 += c2

                sheet.add_chart(c1, 'T'+str(cnt))
                cnt += 10

            sheet = wb['Ê®°ÂûãÂõûÊµã']
            for i in sheet['A1':'L%s' % row_index_8]:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font1'))
                    i[j].alignment = eval(conf.get('config', 'alignment1'))
                    i[j].border = eval(conf.get('config', 'border1'))
            for i in sheet['A1':'L1']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))

            cnt = 2
            while sheet['A'+str(cnt)].value is not None:
                c1 = BarChart()
                v1 = Reference(sheet, min_col=11, min_row=cnt-1, max_col=11, max_row=cnt+9)
                c1.add_data(v1, titles_from_data=True)

                c1.x_axis.title = 'Bin Decile'
                c1.y_axis.title = 'Bad Rate'
                c1.y_axis.majorGridlines = None
                c1.title = sheet['A'+str(cnt)].value

                # Create a second chart
                c2 = LineChart()
                v2 = Reference(sheet, min_col=12, min_row=cnt-1, max_col=12, max_row=cnt+9)
                c2.add_data(v2, titles_from_data=True)
                c2.y_axis.axId = 200
                c2.y_axis.title = "Ks"

                # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
                c1.y_axis.crosses = "max"
                c1 += c2

                sheet.add_chart(c1, 'N'+str(cnt))
                cnt += 10

            wb.create_sheet(title='Êï∞ÊçÆÊµãËØïËØ¥Êòé', index=0)
            head = wb['Êï∞ÊçÆÊµãËØïËØ¥Êòé']
            head['A1'] = 'Êï∞ÊçÆÊµãËØïËØ¥Êòé'
            head.merge_cells('A1:B1')
            head['A2'] = 'Êèê‰æõÁöÑÊ†∑Êú¨'
            head['A3'] = 'ÊµãËØïÊï∞ÊçÆÂèòÈáèÁ±ªÂûã'
            head['A4'] = 'ÊµãËØïÂÜÖÂÆπ'
            head.merge_cells('A4:A9')
            head['A10'] = 'ÊµãËØïÁªìËÆ∫'
            head['B4'] = 'Ê†∑Êú¨ÂàÜÊûê'
            head['B5'] = 'ÂèòÈáèÁº∫Â§±Áéá & Âü∫Êú¨Êé¢Á¥¢ÊÄßÂàÜÊûê'
            head['B6'] = 'INSÂèòÈáèIVÂÄº & Êó∂Èó¥Á®≥ÂÆöÊÄß'
            head['B7'] = 'ÂèòÈáèÈÄæÊúüÂàÜÂ∏ÉÂíåKSÂÄº'
            head['B8'] = 'ËØÑÂàÜÁõ∏ÂÖ≥ÊÄß'
            head['B9'] = 'Ê®°Âûã'
            for i in head['A1':'B10']:
                for j in range(len(i)):
                    if str(i[j])[15:18] in ('B4>', 'B5>', 'B6>', 'B7>', 'B8>', 'B9'):
                        i[j].font = eval(conf.get('config', 'font4'))
                        i[j].alignment = eval(conf.get('config', 'alignment1'))
                        i[j].border = eval(conf.get('config', 'border1'))
                    else:
                        i[j].font = eval(conf.get('config', 'font4'))
                        i[j].alignment = eval(conf.get('config', 'alignment1'))
                        i[j].border = eval(conf.get('config', 'border1'))
            for i in head['A2':'A10']:
                for j in range(len(i)):
                    i[j].font = eval(conf.get('config', 'font2'))
                    i[j].fill = eval(conf.get('config', 'fill1'))
            for i in ['A1', 'B1']:
                head[i].font = eval(conf.get('config', 'font2'))
                head[i].fill = eval(conf.get('config', 'fill1'))

            wb.save(self.file_path)
        else:
            pass
        print('Finish üçµ ')
